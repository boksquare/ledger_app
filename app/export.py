"""Monthly export to Excel (native embedded pie chart) and PDF (rendered chart image)."""
import io

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response

from . import db, queries
from .main import MONTH_RE

router = APIRouter()

SPLIT_LABELS = {"50_50": "50%", "100_hers": "100%"}


def _month_data(month: str) -> dict:
    if not MONTH_RE.match(month):
        raise HTTPException(400, "Month must be YYYY-MM")
    conn = db.get_db()
    try:
        return {
            "expenses": queries.month_expenses(conn, month),
            "summary": queries.month_summary(conn, month),
            "breakdown": queries.category_breakdown(conn, month),
        }
    finally:
        conn.close()


@router.get("/export/{month}.xlsx")
def export_excel(month: str):
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    data = _month_data(month)
    wb = Workbook()

    ws = wb.active
    ws.title = "Expenses"
    ws.append(["Date", "Description", "Category", "Split", "Amount", "Her Share", "Notes"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for e in data["expenses"]:
        her = e["amount"] * 0.5 if e["split_type"] == "50_50" else e["amount"]
        ws.append([e["date"], e["description"], e["category_name"],
                   SPLIT_LABELS[e["split_type"]], e["amount"], round(her, 2), e["notes"]])
    s = data["summary"]
    ws.append([])
    ws.append(["", "", "", "Total spending", s["total_spending"]])
    ws.append(["", "", "", "50/50 total", s["total_50_50"]])
    ws.append(["", "", "", "100% hers total", s["total_100_hers"]])
    ws.append(["", "", "", "HER TOTAL", "", s["her_owed"]])
    for row in ws.iter_rows(min_row=ws.max_row - 3):
        for cell in row:
            cell.font = Font(bold=True)
    for col, width in enumerate([12, 40, 18, 14, 12, 12, 30], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "$#,##0.00"

    chart_ws = wb.create_sheet("Chart")
    chart_ws.append(["Category", "Total"])
    for slice_ in data["breakdown"]:
        chart_ws.append([slice_["name"], slice_["total"]])
    if data["breakdown"]:
        pie = PieChart()
        pie.title = f"Spending by category — {month}"
        n = len(data["breakdown"])
        pie.add_data(Reference(chart_ws, min_col=2, min_row=1, max_row=n + 1), titles_from_data=True)
        pie.set_categories(Reference(chart_ws, min_col=1, min_row=2, max_row=n + 1))
        pie.height, pie.width = 12, 16
        chart_ws.add_chart(pie, "D2")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="expenses-{month}.xlsx"'},
    )


# Same muted qualitative palette as the dashboard chart.
CHART_PALETTE = [
    "#8a9b6e", "#3d5a80", "#c9a227", "#c08497",
    "#b26e4b", "#5f7470", "#85678f", "#d3b88c",
]


def _render_pie_png(breakdown: list[dict]) -> bytes:
    """Donut with a side legend — labels never crowd the slices."""
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    values = [s["total"] for s in breakdown]
    total = sum(values) or 1
    fig, ax = plt.subplots(figsize=(7.4, 3.2))
    wedges, _ = ax.pie(
        values,
        colors=[CHART_PALETTE[i % len(CHART_PALETTE)] for i in range(len(values))],
        startangle=90,
        counterclock=False,
        wedgeprops={"width": 0.42, "edgecolor": "white", "linewidth": 1.5},
    )
    legend_labels = [
        f"{s['name']}   ${s['total']:,.2f}  ({s['total'] / total * 100:.0f}%)"
        for s in breakdown
    ]
    ax.legend(
        wedges, legend_labels,
        loc="center left", bbox_to_anchor=(1.08, 0.5),
        frameon=False, fontsize=9.5, labelcolor="#1c2321", handlelength=1.2,
    )
    ax.set_aspect("equal")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200, bbox_inches="tight", transparent=True)
    plt.close(fig)
    return buf.getvalue()


@router.get("/export/{month}.pdf")
def export_pdf(month: str):
    from datetime import datetime

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable, Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    INK = colors.HexColor("#1c2321")
    MUTED = colors.HexColor("#5c6660")
    ACCENT = colors.HexColor("#3d5a80")
    HIGHLIGHT = colors.HexColor("#bc6c25")
    LINE = colors.HexColor("#e5e2da")
    ROW_ALT = colors.HexColor("#f7f6f3")

    brand = ParagraphStyle("brand", fontName="Helvetica-Bold", fontSize=9,
                           textColor=MUTED, spaceAfter=2)
    title = ParagraphStyle("title", fontName="Times-Bold", fontSize=20,
                           textColor=INK, leading=24, spaceAfter=14)
    label = ParagraphStyle("label", fontName="Helvetica-Bold", fontSize=8,
                           textColor=MUTED, spaceAfter=3)
    big = ParagraphStyle("big", fontName="Times-Bold", fontSize=32,
                         textColor=HIGHLIGHT, leading=36, spaceAfter=4)
    formula = ParagraphStyle("formula", fontName="Helvetica", fontSize=9.5,
                             textColor=MUTED, spaceAfter=6)
    section = ParagraphStyle("section", fontName="Helvetica-Bold", fontSize=11,
                             textColor=INK, spaceBefore=16, spaceAfter=10)
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=8.5,
                          textColor=INK, leading=11)

    data = _month_data(month)
    s = data["summary"]
    month_name = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, title=f"Ledger — {month_name}",
        leftMargin=0.9 * inch, rightMargin=0.9 * inch,
        topMargin=0.8 * inch, bottomMargin=0.8 * inch,
    )

    story = [
        Paragraph("LEDGER", brand),
        Paragraph(f"{month_name} expense report", title),
        Paragraph("HER TOTAL", label),
        Paragraph(f"${s['her_owed']:,.2f}", big),
        Paragraph(
            f"${s['half_share']:,.2f} (50% split)  +  ${s['total_100_hers']:,.2f} (100%)"
            f"   ·   total spending ${s['total_spending']:,.2f}",
            formula,
        ),
        Spacer(1, 6),
        HRFlowable(width="100%", thickness=0.75, color=LINE),
    ]

    if data["breakdown"]:
        story += [
            Paragraph("Spending by category", section),
            Image(io.BytesIO(_render_pie_png(data["breakdown"])),
                  width=6.4 * inch, height=6.4 * inch * (3.2 / 7.4), hAlign="CENTER"),
        ]

    story.append(Paragraph("Itemized expenses", section))
    rows = [["Date", "Description", "Category", "Split", "Amount", "Her Share"]]
    for e in data["expenses"]:
        her = e["amount"] * 0.5 if e["split_type"] == "50_50" else e["amount"]
        rows.append([
            e["date"], Paragraph(e["description"], body), e["category_name"],
            SPLIT_LABELS[e["split_type"]], f"${e['amount']:,.2f}", f"${her:,.2f}",
        ])
    rows.append(["", "", "", "", "Her total", f"${s['her_owed']:,.2f}"])
    table = Table(
        rows,
        colWidths=[0.85 * inch, 2.6 * inch, 1.15 * inch, 0.6 * inch, 0.85 * inch, 0.85 * inch],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        # header
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        # body
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("TEXTCOLOR", (0, 1), (-1, -2), INK),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, ROW_ALT]),
        ("LINEBELOW", (0, 1), (-1, -2), 0.4, LINE),
        # totals row
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, -1), (-1, -1), HIGHLIGHT),
        ("LINEABOVE", (0, -1), (-1, -1), 0.75, ACCENT),
        # shared
        ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    doc.build(story)

    return Response(
        buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="expenses-{month}.pdf"'},
    )
