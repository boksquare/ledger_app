"""Best-effort statement parsing without AI — fallback when Claude is unavailable.

Handles the common shapes: CSV/Excel exports with a header row (Date/Description/
Amount or Debit/Credit columns), headerless positional exports, and text-based PDF
statements whose transaction lines start with a date and end with an amount.
Category suggestions are keyword heuristics and always reviewable before import.
"""
import csv
import io
import re
from datetime import date, datetime
from pathlib import Path

PAYMENT_RE = re.compile(
    r"payment\s+(received|thank)|thank\s+you|autopay|online payment|mobile payment",
    re.I,
)

PDF_LINE_RE = re.compile(
    r"^\s*(?P<date>\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?|\d{4}-\d{2}-\d{2})\s+"
    r"(?P<desc>.*?)\s+"
    r"(?P<amt>\(?-?\$?\d[\d,]*\.\d{2}\)?)(?:\s*(?:CR|DR))?\s*$"
)

# Only applied when a matching category name exists; always overridable in review.
CATEGORY_KEYWORDS = [
    ("Groceries", ["costco", "h-e-b", "heb ", "kroger", "whole foods", "aldi",
                   "trader joe", "safeway", "grocery", "market"]),
    ("Dining", ["uber eats", "grubhub", "doordash", "restaurant", "chipotle",
                "mcdonald", "starbucks", "cafe", "pizza", "taco", "sushi",
                "grill", "kitchen", "bbq", "diner"]),
    ("Subscriptions", ["netflix", "spotify", "hulu", "disney+", "youtube premium",
                       "apple.com/bill", "subscription", "patreon", "prime video"]),
    ("Transportation", ["shell", "exxon", "chevron", "valero", "uber trip",
                        "lyft", "fuel", "parking", "toll", "metro"]),
    ("Internet", ["comcast", "xfinity", "spectrum", "internet", "fios"]),
    ("Utilities", ["electric", "energy", "water", "utility", "utilities"]),
    ("Household", ["home depot", "lowe's", "lowes", "ikea", "target", "amazon"]),
]


def _suggest_category(description: str, category_names: list[str]) -> str | None:
    by_lower = {c.lower(): c for c in category_names}
    d = description.lower()
    for cat, keywords in CATEGORY_KEYWORDS:
        if cat.lower() in by_lower and any(k in d for k in keywords):
            return by_lower[cat.lower()]
    return None


def _parse_date(raw: str, today: date | None = None) -> date | None:
    raw = str(raw).strip()
    today = today or date.today()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y",
                "%b %d, %Y", "%B %d, %Y", "%d %b %Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(raw).date()
    except ValueError:
        pass
    # Month/day with no year (common on statement lines): assume the most
    # recent occurrence that isn't in the future.
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})$", raw)
    if m:
        try:
            d = date(today.year, int(m[1]), int(m[2]))
        except ValueError:
            return None
        return d.replace(year=today.year - 1) if d > today else d
    return None


def _parse_amount(raw: str) -> float | None:
    s = str(raw).strip()
    if not s:
        return None
    negative = (s.startswith("(") and s.endswith(")")) or bool(re.search(r"\bCR\b", s, re.I))
    s = re.sub(r"\(|\)|\$|,|\s|CR|DR", "", s, flags=re.I)
    if not re.fullmatch(r"-?\d+(\.\d{1,2})?", s):
        return None
    value = float(s)
    return round(-abs(value) if negative else value, 2)


def _rows_from_file(path: Path, text: str) -> list[list[str]] | None:
    """Tabular rows for CSV/Excel; None for PDFs (parsed line-by-line instead)."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return [row for row in csv.reader(io.StringIO(text))]
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        rows = []
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rows.append([
                        "" if v is None else (v.isoformat() if hasattr(v, "isoformat") else str(v))
                        for v in row
                    ])
        finally:
            wb.close()
        return rows
    return None


def _find_header(rows: list[list[str]]) -> tuple[int, dict] | None:
    for i, row in enumerate(rows[:15]):
        low = [str(c).strip().lower() for c in row]
        has_date = any("date" in c for c in low)
        has_amount = any("amount" in c or c in ("debit", "credit") for c in low)
        if not (has_date and has_amount):
            continue
        cols: dict = {"date": None, "desc": None, "amount": None, "debit": None, "credit": None}
        for j, c in enumerate(low):
            if cols["date"] is None and "date" in c:
                cols["date"] = j
            if cols["desc"] is None and any(k in c for k in
                    ("description", "merchant", "payee", "details", "memo", "name")):
                cols["desc"] = j
            if cols["amount"] is None and "amount" in c:
                cols["amount"] = j
            if c == "debit":
                cols["debit"] = j
            if c == "credit":
                cols["credit"] = j
        if cols["date"] is not None:
            return i, cols
    return None


def _cell(cells: list[str], idx: int | None) -> str:
    return cells[idx].strip() if idx is not None and idx < len(cells) else ""


def _parse_tabular(rows: list[list[str]], category_names: list[str]) -> list[dict]:
    header = _find_header(rows)
    txns = []
    body = rows[header[0] + 1:] if header else rows
    for row in body:
        cells = [str(c) for c in row]
        if not any(c.strip() for c in cells):
            continue
        if header:
            cols = header[1]
            d = _parse_date(_cell(cells, cols["date"]))
            desc = _cell(cells, cols["desc"])
            amount = _parse_amount(_cell(cells, cols["amount"]))
            if amount is None and cols["debit"] is not None:
                debit = _parse_amount(_cell(cells, cols["debit"]))
                credit = _parse_amount(_cell(cells, cols["credit"]))
                amount = debit if debit is not None else (-credit if credit is not None else None)
        else:
            # No header row: first date-like cell, last amount-like cell,
            # description is whatever sits between them.
            d = date_idx = None
            for j, c in enumerate(cells):
                d = _parse_date(c.strip())
                if d:
                    date_idx = j
                    break
            if d is None:
                continue
            amount = amount_idx = None
            for j in range(len(cells) - 1, date_idx, -1):
                amount = _parse_amount(cells[j].strip())
                if amount is not None:
                    amount_idx = j
                    break
            desc = " ".join(c.strip() for c in cells[date_idx + 1:amount_idx] if c.strip())
        if d is None or amount is None or not desc or PAYMENT_RE.search(desc):
            continue
        txns.append({
            "date": d.isoformat(),
            "description": re.sub(r"\s{2,}", " ", desc).strip(),
            "amount": amount,
            "suggested_category": _suggest_category(desc, category_names),
        })
    return txns


def _parse_pdf_lines(text: str, category_names: list[str]) -> list[dict]:
    txns = []
    for line in text.splitlines():
        m = PDF_LINE_RE.match(line)
        if not m:
            continue
        d = _parse_date(m["date"])
        amount = _parse_amount(m["amt"])
        desc = re.sub(r"\s{2,}", " ", m["desc"]).strip()
        if d is None or amount is None or not desc or PAYMENT_RE.search(desc):
            continue
        txns.append({
            "date": d.isoformat(),
            "description": desc,
            "amount": amount,
            "suggested_category": _suggest_category(desc, category_names),
        })
    return txns


def parse_statement_native(path: Path, text: str, category_names: list[str]) -> dict:
    """Same output shape as the Claude parser. Raises ValueError if nothing parses."""
    rows = _rows_from_file(path, text)
    txns = (_parse_tabular(rows, category_names) if rows is not None
            else _parse_pdf_lines(text, category_names))
    if not txns:
        raise ValueError(
            "the built-in parser couldn't recognize any transactions in this file"
        )
    dates = sorted(t["date"] for t in txns)
    return {
        "card_name": None,
        "period_start": dates[0],
        "period_end": dates[-1],
        "transactions": txns,
    }
