"""Statement upload, parsing, staging/review, and confirm-import (spec §6 + §9)."""
import re
import time
from pathlib import Path

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from starlette.datastructures import UploadFile

from . import db, queries
from .claude_parser import ClaudeParsingError, parse_statement_text
from .native_parser import parse_statement_native

router = APIRouter()

ALLOWED_EXTENSIONS = {".pdf", ".csv", ".xlsx", ".xlsm", ".xls"}
# Heuristic: a text-based PDF yields far more than this per page; below it,
# the file is likely a scanned image (OCR is out of scope for v1).
MIN_CHARS_PER_PAGE = 100
MAX_EXCEL_ROWS = 5000


def _templates():
    from .main import templates
    return templates


def _excel_to_text(path: Path) -> str:
    """Flatten every sheet into tab-separated lines for Claude to parse."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise ValueError("This Excel file could not be opened — is it a valid .xlsx?")
    lines: list[str] = []
    try:
        for ws in wb.worksheets:
            if len(wb.worksheets) > 1:
                lines.append(f"=== Sheet: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [
                    "" if v is None else (v.isoformat() if hasattr(v, "isoformat") else str(v))
                    for v in row
                ]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells))
                if len(lines) > MAX_EXCEL_ROWS:
                    raise ValueError(
                        f"This spreadsheet has more than {MAX_EXCEL_ROWS} rows — "
                        "export just the statement period and try again."
                    )
    finally:
        wb.close()
    if not lines:
        raise ValueError("This Excel file appears to be empty.")
    return "\n".join(lines)


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return path.read_text(errors="replace")
    if suffix in (".xlsx", ".xlsm"):
        return _excel_to_text(path)
    if suffix == ".xls":
        raise ValueError(
            "Legacy .xls files aren't supported — open it in Excel and save as .xlsx, "
            "then upload again."
        )
    import pdfplumber

    with pdfplumber.open(path) as pdf:
        pages = [page.extract_text() or "" for page in pdf.pages]
    text = "\n".join(pages)
    if len(text.strip()) < MIN_CHARS_PER_PAGE * max(len(pages), 1):
        raise ValueError(
            "This looks like a scanned PDF (little or no selectable text). "
            "OCR is not supported yet — try downloading a text-based statement from your bank."
        )
    return text


@router.get("/import", response_class=HTMLResponse)
def import_page(request: Request, error: str = "", ok: str = "", warn: str = ""):
    conn = db.get_db()
    try:
        imports = conn.execute(
            """
            SELECT si.*,
                SUM(CASE WHEN st.status = 'pending' THEN 1 ELSE 0 END) AS pending_count,
                SUM(CASE WHEN st.status = 'imported' THEN 1 ELSE 0 END) AS imported_count,
                SUM(CASE WHEN st.status = 'discarded' THEN 1 ELSE 0 END) AS discarded_count
            FROM statement_imports si
            LEFT JOIN staged_transactions st ON st.statement_import_id = si.id
            GROUP BY si.id ORDER BY si.uploaded_at DESC
            """
        ).fetchall()
    finally:
        conn.close()
    return _templates().TemplateResponse(
        request, "import.html",
        {"imports": imports, "error": error, "ok": ok, "warn": warn},
    )


@router.post("/import/upload")
async def upload_statements(request: Request):
    form = await request.form()
    files = [f for f in form.getlist("statements") if isinstance(f, UploadFile) and f.filename]
    if not files:
        return RedirectResponse("/import?error=No+files+selected", status_code=303)

    created_ids, errors, warnings = [], [], []
    conn = db.get_db()
    try:
        cats = queries.active_categories(conn)
        cat_names = [c["name"] for c in cats]
        cat_by_name = {c["name"].lower(): c["id"] for c in cats}

        for upload in files:
            ext = Path(upload.filename).suffix.lower()
            if ext not in ALLOWED_EXTENSIONS:
                errors.append(f"{upload.filename}: only PDF, CSV, and Excel files are supported")
                continue
            safe_name = re.sub(r"[^\w.\-]", "_", upload.filename)
            dest = db.STATEMENTS_DIR / f"{int(time.time())}_{safe_name}"
            dest.write_bytes(await upload.read())

            try:
                text = extract_text(dest)
            except ValueError as e:
                errors.append(f"{upload.filename}: {e}")
                dest.unlink(missing_ok=True)
                continue
            try:
                parsed = parse_statement_text(text, cat_names)
            except ClaudeParsingError as claude_err:
                # AI unavailable or failed — fall back to the built-in parser.
                try:
                    parsed = parse_statement_native(dest, text, cat_names)
                    warnings.append(
                        f"{upload.filename}: Claude was unavailable, so the built-in "
                        f"parser was used instead — double-check dates, amounts, and "
                        f"categories before confirming. (Claude error: {claude_err})"
                    )
                except ValueError as native_err:
                    errors.append(
                        f"{upload.filename}: {claude_err} The built-in fallback "
                        f"couldn't parse it either ({native_err})."
                    )
                    dest.unlink(missing_ok=True)
                    continue

            cur = conn.execute(
                """
                INSERT INTO statement_imports
                    (uploaded_filename, card_name, statement_period_start,
                     statement_period_end, original_file_path)
                VALUES (?, ?, ?, ?, ?)
                """,
                (upload.filename, parsed.get("card_name"), parsed.get("period_start"),
                 parsed.get("period_end"), str(dest)),
            )
            import_id = cur.lastrowid
            for txn in parsed.get("transactions", []):
                suggested = (txn.get("suggested_category") or "").lower()
                conn.execute(
                    """
                    INSERT INTO staged_transactions
                        (statement_import_id, date, description, amount, suggested_category_id)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (import_id, txn["date"], txn["description"], txn["amount"],
                     cat_by_name.get(suggested)),
                )
            conn.commit()
            created_ids.append(import_id)
    finally:
        conn.close()

    from urllib.parse import urlencode
    params = {}
    if errors:
        params["error"] = "; ".join(errors)
    if warnings:
        params["warn"] = "; ".join(warnings)
    query = f"?{urlencode(params)}" if params else ""
    if created_ids:
        return RedirectResponse(f"/import/{created_ids[0]}{query}", status_code=303)
    return RedirectResponse(f"/import{query}", status_code=303)


@router.get("/import/{import_id}", response_class=HTMLResponse)
def review_page(request: Request, import_id: int, error: str = "", ok: str = "", warn: str = ""):
    conn = db.get_db()
    try:
        imp = conn.execute(
            "SELECT * FROM statement_imports WHERE id = ?", (import_id,)
        ).fetchone()
        if not imp:
            raise HTTPException(404, "Statement not found")
        txns = conn.execute(
            """
            SELECT st.*, c.name AS suggested_category_name
            FROM staged_transactions st
            LEFT JOIN categories c ON c.id = st.suggested_category_id
            WHERE st.statement_import_id = ?
            ORDER BY st.date, st.id
            """,
            (import_id,),
        ).fetchall()
        cats = queries.active_categories(conn)
    finally:
        conn.close()
    return _templates().TemplateResponse(
        request, "import_review.html",
        {"imp": imp, "txns": txns, "categories": cats,
         "error": error, "ok": ok, "warn": warn},
    )


@router.post("/import/{import_id}/confirm")
async def confirm_import(request: Request, import_id: int):
    form = await request.form()
    discard_unchecked = form.get("discard_unchecked") == "on"
    imported = skipped = 0

    conn = db.get_db()
    try:
        pending = conn.execute(
            "SELECT * FROM staged_transactions WHERE statement_import_id = ? AND status = 'pending'",
            (import_id,),
        ).fetchall()
        if not pending:
            raise HTTPException(400, "No pending transactions for this statement")

        for txn in pending:
            tid = txn["id"]
            if form.get(f"selected_{tid}") != "on":
                if discard_unchecked:
                    conn.execute(
                        "UPDATE staged_transactions SET status = 'discarded' WHERE id = ?", (tid,)
                    )
                continue
            try:
                amount = round(float(form.get(f"amount_{tid}", txn["amount"])), 2)
            except ValueError:
                amount = 0
            if amount <= 0:
                skipped += 1
                continue
            date = form.get(f"date_{tid}") or txn["date"]
            description = (form.get(f"description_{tid}") or txn["description"]).strip()
            category_id = int(form.get(f"category_{tid}"))
            split_type = form.get(f"split_{tid}", "50_50")
            if split_type not in db.SPLIT_TYPES:
                split_type = "50_50"
            conn.execute(
                """
                INSERT INTO expenses
                    (date, description, amount, category_id, split_type, statement_import_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (date, description, amount, category_id, split_type, import_id),
            )
            conn.execute(
                "UPDATE staged_transactions SET status = 'imported' WHERE id = ?", (tid,)
            )
            imported += 1
        conn.commit()
    finally:
        conn.close()

    msg = f"Imported {imported} expense{'s' if imported != 1 else ''}"
    if skipped:
        msg += f" ({skipped} skipped: amount must be positive)"
    from urllib.parse import quote
    return RedirectResponse(f"/import/{import_id}?ok={quote(msg)}", status_code=303)
